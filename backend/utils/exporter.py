"""
PCForge AI — Export Utility
Generates CSV and Excel (.xlsx) exports from AnalyzeResponse data.
Excel workbook contains 6 structured sheets.
"""
from __future__ import annotations
import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Fill, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from backend.models.schemas import AnalyzeResponse

logger = logging.getLogger(__name__)

# ─── Color Palette ────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
_ACCENT_FILL = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
_LIVE_FILL = PatternFill(start_color="0D3B1E", end_color="0D3B1E", fill_type="solid")
_PRED_FILL = PatternFill(start_color="3B1F0D", end_color="3B1F0D", fill_type="solid")
_ERROR_FILL = PatternFill(start_color="3B0D0D", end_color="3B0D0D", fill_type="solid")
_WARN_FILL = PatternFill(start_color="3B2E0D", end_color="3B2E0D", fill_type="solid")
_PASS_FILL = PatternFill(start_color="0D3B0D", end_color="0D3B0D", fill_type="solid")

_HEADER_FONT = Font(name="Calibri", bold=True, color="E0E0E0", size=11)
_BODY_FONT = Font(name="Calibri", color="DCDCDC", size=10)
_TITLE_FONT = Font(name="Calibri", bold=True, color="7EB8FF", size=13)
_WARN_FONT = Font(name="Calibri", color="FFD580", size=10)
_ERROR_FONT = Font(name="Calibri", color="FF6B6B", size=10)
_GREEN_FONT = Font(name="Calibri", color="6BFF9C", size=10)

_THIN_BORDER = Border(
    left=Side(style="thin", color="333355"),
    right=Side(style="thin", color="333355"),
    top=Side(style="thin", color="333355"),
    bottom=Side(style="thin", color="333355"),
)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _style_body_row(ws, row: int, n_cols: int, fill=None) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill or _ACCENT_FILL
        cell.font = _BODY_FONT
        cell.alignment = _LEFT
        cell.border = _THIN_BORDER


def _auto_col_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _add_title(ws, title: str, row: int = 1) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.fill = PatternFill(start_color="0F0F1E", end_color="0F0F1E", fill_type="solid")
    cell.alignment = _LEFT


# ─── Sheet 1: User Input ──────────────────────────────────────────────────────

def _sheet_user_input(wb: Workbook, analysis: AnalyzeResponse) -> None:
    ws = wb.create_sheet("User Input")
    ws.sheet_properties.tabColor = "4A90D9"

    _add_title(ws, "📋 User Input — Build Specification", row=1)
    ws.merge_cells("A1:B1")

    headers = ["Field", "Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    raw = analysis.original_input
    rows = [
        ("CPU", raw.get("cpu", "Not provided")),
        ("GPU", raw.get("gpu", "Not provided")),
        ("Motherboard", raw.get("motherboard", "Not provided")),
        ("RAM", str(raw.get("ram", "Not provided"))),
        ("Storage", str(raw.get("storage", "Not provided"))),
        ("PSU", raw.get("psu", "Not provided")),
        ("Case", raw.get("case", "Not provided")),
        ("Cooler", raw.get("cooler", "Not provided")),
        ("Monitor", raw.get("monitor", "Not provided")),
        ("Budget (USD)", f"${raw['budget_usd']:.2f}" if raw.get("budget_usd") else "Not specified"),
        ("Usage Type", raw.get("usage_type", "Not specified")),
        ("Region", raw.get("region", "US")),
        ("Preferred Brand", raw.get("preferred_brand", "None")),
        ("Build ID", analysis.build_id),
        ("Analysis Time", analysis.timestamp.isoformat()),
        ("Inferred Tier", analysis.inferred_tier),
    ]

    for r_idx, (field, value) in enumerate(rows, 3):
        ws.cell(row=r_idx, column=1, value=field)
        ws.cell(row=r_idx, column=2, value=str(value))
        _style_body_row(ws, r_idx, 2)

    _auto_col_width(ws)


# ─── Sheet 2: Final Build ─────────────────────────────────────────────────────

def _sheet_final_build(wb: Workbook, analysis: AnalyzeResponse) -> None:
    ws = wb.create_sheet("Final Build")
    ws.sheet_properties.tabColor = "27AE60"

    _add_title(ws, "🖥️ Final Build — Completed Component List", row=1)
    ws.merge_cells("A1:D1")

    headers = ["Category", "Brand", "Model", "Auto-Filled?"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    for r_idx, comp in enumerate(analysis.completed_build, 3):
        ws.cell(row=r_idx, column=1, value=comp.category)
        ws.cell(row=r_idx, column=2, value=comp.brand)
        ws.cell(row=r_idx, column=3, value=comp.model)
        auto_filled = "✅ Auto-filled" if comp.is_auto_filled else "User provided"
        ws.cell(row=r_idx, column=4, value=auto_filled)
        fill = _LIVE_FILL if not comp.is_auto_filled else _ACCENT_FILL
        _style_body_row(ws, r_idx, 4, fill=fill)
        if comp.is_auto_filled:
            ws.cell(row=r_idx, column=4).font = _WARN_FONT

    _auto_col_width(ws)


# ─── Sheet 3: Price Breakdown ─────────────────────────────────────────────────

def _sheet_price_breakdown(wb: Workbook, analysis: AnalyzeResponse) -> None:
    ws = wb.create_sheet("Price Breakdown")
    ws.sheet_properties.tabColor = "F39C12"

    _add_title(ws, "💰 Price Breakdown — Part-wise Pricing", row=1)
    ws.merge_cells("A1:H1")

    headers = ["Category", "Brand", "Model", "Price (USD)", "Source", "Store", "Availability", "URL"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    total = 0.0
    for r_idx, part in enumerate(analysis.pricing, 3):
        ws.cell(row=r_idx, column=1, value=part.category)
        ws.cell(row=r_idx, column=2, value=part.brand)
        ws.cell(row=r_idx, column=3, value=part.model)
        ws.cell(row=r_idx, column=4, value=f"${part.price_usd:.2f}")
        ws.cell(row=r_idx, column=5, value=part.source.upper())
        ws.cell(row=r_idx, column=6, value=part.store)
        ws.cell(row=r_idx, column=7, value=part.availability)
        ws.cell(row=r_idx, column=8, value=part.url)
        total += part.price_usd

        # Color-code by source
        fill = _LIVE_FILL if part.source in ("live", "simulated") else _PRED_FILL
        _style_body_row(ws, r_idx, 8, fill=fill)

        # Source label font
        source_cell = ws.cell(row=r_idx, column=5)
        if part.source == "predicted":
            source_cell.font = _WARN_FONT
        else:
            source_cell.font = _GREEN_FONT

    # Total row
    total_row = len(analysis.pricing) + 3
    ws.cell(row=total_row, column=3, value="TOTAL")
    ws.cell(row=total_row, column=4, value=f"${total:.2f}")
    for c in range(1, 9):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER

    # Market range
    mr = analysis.price_summary.market_range
    range_row = total_row + 2
    ws.cell(row=range_row, column=1, value="Market Range")
    ws.cell(row=range_row, column=2, value=f"Min: ${mr.min_price:.2f}")
    ws.cell(row=range_row, column=3, value=f"Avg: ${mr.average_price:.2f}")
    ws.cell(row=range_row, column=4, value=f"Max: ${mr.max_price:.2f}")
    _style_header_row(ws, range_row, 4)

    _auto_col_width(ws)


# ─── Sheet 4: Recommendations ─────────────────────────────────────────────────

def _sheet_recommendations(wb: Workbook, analysis: AnalyzeResponse) -> None:
    ws = wb.create_sheet("Recommendations")
    ws.sheet_properties.tabColor = "8E44AD"

    _add_title(ws, f"🎯 Recommendations — {analysis.recommendations.inferred_tier.title()} Build", row=1)
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value=f"Tier Reasoning: {analysis.recommendations.tier_reasoning}")
    ws.cell(row=2, column=1).font = _BODY_FONT
    ws.merge_cells("A2:D2")

    # Primary recommendations
    headers = ["Category", "Model", "Price (USD)", "Reasoning"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, len(headers))

    row = 4
    for rec in analysis.recommendations.recommended_parts:
        ws.cell(row=row, column=1, value=rec.category)
        ws.cell(row=row, column=2, value=rec.model)
        ws.cell(row=row, column=3, value=f"${rec.price_usd:.2f}")
        ws.cell(row=row, column=4, value=rec.reasoning)
        _style_body_row(ws, row, 4)
        row += 1

    # Alternatives
    row += 1
    ws.cell(row=row, column=1, value="── Alternative Options ──")
    ws.cell(row=row, column=1).font = _TITLE_FONT
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    alt_headers = ["Category", "Alternative Model", "Price (USD)", "Notes"]
    for c, h in enumerate(alt_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, 4)
    row += 1

    for category, alts in analysis.recommendations.alternatives.items():
        for alt in alts:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=alt.model)
            ws.cell(row=row, column=3, value=f"${alt.price_usd:.2f}")
            ws.cell(row=row, column=4, value=alt.notes)
            _style_body_row(ws, row, 4, fill=_ACCENT_FILL)
            row += 1

    _auto_col_width(ws)


# ─── Sheet 5: Compatibility Report ───────────────────────────────────────────

def _sheet_compatibility(wb: Workbook, analysis: AnalyzeResponse) -> None:
    ws = wb.create_sheet("Compatibility Report")
    ws.sheet_properties.tabColor = "E74C3C"

    compat = analysis.compatibility
    status_colors = {"valid": "✅ VALID", "warning": "⚠️ WARNING", "invalid": "❌ INVALID"}
    _add_title(ws, f"🔧 Compatibility Report — {status_colors.get(compat.status, compat.status)}", row=1)
    ws.merge_cells("A1:D1")

    # Summary stats
    ws.cell(row=2, column=1, value=f"Total Checks: {compat.total_checks}")
    ws.cell(row=2, column=2, value=f"Issues Found: {len(compat.issues)}")
    ws.cell(row=2, column=3, value=f"Passed: {len(compat.passed_checks)}")
    for c in range(1, 4):
        ws.cell(row=2, column=c).font = _HEADER_FONT
        ws.cell(row=2, column=c).fill = _HEADER_FILL

    # Passed checks
    row = 4
    ws.cell(row=row, column=1, value="Passed Checks")
    ws.cell(row=row, column=1).font = _TITLE_FONT
    row += 1
    for check in compat.passed_checks:
        ws.cell(row=row, column=1, value=f"  ✓ {check}")
        ws.cell(row=row, column=1).font = _GREEN_FONT
        ws.cell(row=row, column=1).fill = _PASS_FILL
        row += 1

    # Issues
    if compat.issues:
        row += 1
        headers = ["Severity", "Component", "Issue", "Suggested Fix"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 4)
        row += 1

        for issue in compat.issues:
            ws.cell(row=row, column=1, value=issue.severity.upper())
            ws.cell(row=row, column=2, value=issue.component)
            ws.cell(row=row, column=3, value=issue.issue)
            ws.cell(row=row, column=4, value=issue.suggested_fix)

            if issue.severity == "error":
                fill = _ERROR_FILL
                font = _ERROR_FONT
            elif issue.severity == "warning":
                fill = _WARN_FILL
                font = _WARN_FONT
            else:
                fill = _ACCENT_FILL
                font = _BODY_FONT

            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = fill
                ws.cell(row=row, column=c).font = font
                ws.cell(row=row, column=c).border = _THIN_BORDER
                ws.cell(row=row, column=c).alignment = _LEFT

            row += 1

    _auto_col_width(ws)


# ─── Sheet 6: Metadata ────────────────────────────────────────────────────────

def _sheet_metadata(wb: Workbook, analysis: AnalyzeResponse) -> None:
    ws = wb.create_sheet("Metadata")
    ws.sheet_properties.tabColor = "2C3E50"

    _add_title(ws, "📊 Metadata — Analysis Provenance", row=1)
    ws.merge_cells("A1:B1")

    headers = ["Key", "Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, 2)

    summary = analysis.price_summary
    live_count = summary.live_parts_count
    pred_count = summary.predicted_parts_count

    metadata_rows = [
        ("Build ID", analysis.build_id),
        ("Analysis Timestamp", analysis.timestamp.isoformat()),
        ("Export Generated At", datetime.now(timezone.utc).isoformat()),
        ("Inferred Build Tier", analysis.inferred_tier),
        ("Usage Type", analysis.usage_type or "Not specified"),
        ("Region", summary.region),
        ("Currency", summary.currency),
        ("─── Pricing ───", ""),
        ("Total Live/Simulated Price", f"${summary.total_live_usd:.2f}"),
        ("Total Predicted Price", f"${summary.total_predicted_usd:.2f}"),
        ("Total Combined Price", f"${summary.total_combined_usd:.2f}"),
        ("Market Price Range (Min)", f"${summary.market_range.min_price:.2f}"),
        ("Market Price Range (Avg)", f"${summary.market_range.average_price:.2f}"),
        ("Market Price Range (Max)", f"${summary.market_range.max_price:.2f}"),
        ("Parts with Live/Simulated Price", str(live_count)),
        ("Parts with Predicted Price", str(pred_count)),
        ("─── Flags ───", ""),
        ("Compatibility Status", analysis.compatibility.status),
        ("Auto-Filled Components", ", ".join(analysis.auto_filled_components) or "None"),
        ("─── Data Sources ───", ""),
        ("Live Price Source", "Simulated market data (PCForge AI v1.0)"),
        ("ML Model", "XGBoost Regressor / Heuristic Fallback"),
        ("Prediction Accuracy", "±10–15% of real market prices"),
        ("Price Data As Of", analysis.timestamp.strftime("%Y-%m-%d")),
        ("─── Notes ───", ""),
    ]

    for note in analysis.notes:
        metadata_rows.append(("Note", note))

    for r_idx, (key, value) in enumerate(metadata_rows, 3):
        ws.cell(row=r_idx, column=1, value=key)
        ws.cell(row=r_idx, column=2, value=str(value))
        fill = _HEADER_FILL if key.startswith("─") else _ACCENT_FILL
        font = _TITLE_FONT if key.startswith("─") else _BODY_FONT
        for c in range(1, 3):
            cell = ws.cell(row=r_idx, column=c)
            cell.fill = fill
            cell.font = font
            cell.border = _THIN_BORDER
            cell.alignment = _LEFT

    _auto_col_width(ws)


# ─── Main Export Functions ────────────────────────────────────────────────────

def export_excel(analysis: AnalyzeResponse) -> bytes:
    """
    Generate full Excel workbook with 6 sheets.
    Returns bytes for streaming response.
    """
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_user_input(wb, analysis)
    _sheet_final_build(wb, analysis)
    _sheet_price_breakdown(wb, analysis)
    _sheet_recommendations(wb, analysis)
    _sheet_compatibility(wb, analysis)
    _sheet_metadata(wb, analysis)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_csv(analysis: AnalyzeResponse) -> str:
    """
    Generate flat CSV export of the price breakdown.
    Returns CSV string.
    """
    rows = []
    for part in analysis.pricing:
        rows.append({
            "build_id": analysis.build_id,
            "timestamp": analysis.timestamp.isoformat(),
            "category": part.category,
            "brand": part.brand,
            "model": part.model,
            "price_usd": part.price_usd,
            "source": part.source,
            "store": part.store,
            "availability": part.availability,
            "url": part.url,
            "is_auto_filled": part.category in analysis.auto_filled_components,
            "compatibility_status": analysis.compatibility.status,
            "inferred_tier": analysis.inferred_tier,
        })

    if not rows:
        return ""

    df = pd.DataFrame(rows)
    # Append summary rows
    summary_rows = pd.DataFrame([
        {"category": "─ TOTAL (Live) ─", "price_usd": analysis.price_summary.total_live_usd},
        {"category": "─ TOTAL (Predicted) ─", "price_usd": analysis.price_summary.total_predicted_usd},
        {"category": "─ TOTAL (Combined) ─", "price_usd": analysis.price_summary.total_combined_usd},
        {"category": "─ Market Min ─", "price_usd": analysis.price_summary.market_range.min_price},
        {"category": "─ Market Avg ─", "price_usd": analysis.price_summary.market_range.average_price},
        {"category": "─ Market Max ─", "price_usd": analysis.price_summary.market_range.max_price},
    ])
    df = pd.concat([df, summary_rows], ignore_index=True)
    return df.to_csv(index=False)
